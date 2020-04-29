# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import pandas as pd
import os
from openpyxl import load_workbook

cur_path = os.getcwd()
up_path = os.path.dirname(cur_path)


class ExcelExporter(object):
    def process_item(self, item, spider):
        file_path = os.path.join(up_path, '{}.xlsx'.format(spider.name))
        is_exist = os.path.exists(file_path)
        columns = item.keys()
        df = pd.DataFrame([item], columns=columns)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')

        if not is_exist:
            df.to_excel(writer, header=True, sheet_name=spider.name, index=False, columns=columns)
            writer.save()

        else:
            book = load_workbook(file_path)
            row_start = book[spider.name].max_row
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, header=False, sheet_name=spider.name, index=False, columns=columns,
                        startrow=row_start + 1)
            writer.save()

        return item


class CsvExporter(object):
    def process_item(self, item, spider):
        file_path = os.path.join(up_path, '{}.csv'.format(spider.name))
        is_exist = os.path.exists(file_path)
        columns = item.keys()
        df = pd.DataFrame([item], columns=columns)
        with open(file_path, mode='a') as output_file:
            if not is_exist:
                df.to_csv(output_file, index=False, header=True, encoding='GBK')
            else:
                df.to_csv(output_file, index=False, header=False, encoding='GBK')
        return item
